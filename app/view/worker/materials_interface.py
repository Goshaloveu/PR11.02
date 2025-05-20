from PyQt6.QtCore import Qt, pyqtSlot
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFormLayout, QLabel, 
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView
)
from PyQt6.QtGui import QIcon, QColor, QFont
import re
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import os
from datetime import datetime

from qfluentwidgets import (
    ScrollArea, LineEdit, PushButton, PrimaryPushButton, 
    InfoBar, MessageBox, SubtitleLabel, BodyLabel,
    CardWidget, FluentIcon, StrongBodyLabel, ExpandLayout,
    Dialog, SpinBox, ComboBox, TableWidget, FlowLayout,
    PushButton, SearchLineEdit, TextEdit, IndeterminateProgressBar
)

from ...common.db.database import SessionLocal
from ...common.db.controller import MaterialController
from ...common.db.models_pydantic import Material, MaterialCreate, MaterialUpdate
from ...common.signal_bus import signalBus


class AddMaterialDialog(MessageBox):
    def __init__(self, parent=None):
        # Initialize with title and placeholder content
        super().__init__(
            title="Добавление материала", 
            content="", 
            parent=parent
        )
        
        # Replace content with our widget
        self.contentLabel.hide()
        
        # Create custom content
        self.form_widget = QWidget()
        form_layout = QVBoxLayout(self.form_widget)
        
        # Form layout
        fields_layout = QFormLayout()
        fields_layout.setVerticalSpacing(10)
        fields_layout.setHorizontalSpacing(20)
        
        # Add fields
        self.type_input = LineEdit()
        self.type_input.setPlaceholderText("Тип материала")
        
        self.balance_input = SpinBox()
        self.balance_input.setRange(0, 999999)
        self.balance_input.setValue(0)
        
        self.price_input = SpinBox()
        self.price_input.setRange(1, 999999)
        self.price_input.setValue(100)
        self.price_input.setSuffix(" ₽")
        
        # Add fields to form
        fields_layout.addRow("Тип материала:", self.type_input)
        fields_layout.addRow("Начальный остаток:", self.balance_input)
        fields_layout.addRow("Цена за единицу:", self.price_input)
        
        form_layout.addLayout(fields_layout)
        
        # Insert our form in place of content label
        self.textLayout.insertWidget(2, self.form_widget)
        
        # Update dialog size
        self.widget.setMinimumWidth(400)
        self.widget.setMinimumHeight(300)
        
        # Update button text
        self.yesButton.setText("Добавить")
        self.cancelButton.setText("Отмена")
        
    def get_material_data(self):
        return {
            "type": self.type_input.text(),
            "balance": self.balance_input.value(),
            "price": self.price_input.value()
        }


class EditMaterialDialog(MessageBox):
    def __init__(self, material, parent=None):
        # Initialize with title and placeholder content
        super().__init__(
            title="Редактирование материала", 
            content="", 
            parent=parent
        )
        
        self.material = material
        
        # Replace content with our widget
        self.contentLabel.hide()
        
        # Create custom content
        self.form_widget = QWidget()
        form_layout = QVBoxLayout(self.form_widget)
        
        # Form layout
        fields_layout = QFormLayout()
        fields_layout.setVerticalSpacing(10)
        fields_layout.setHorizontalSpacing(20)
        
        # Add fields
        self.type_input = LineEdit()
        self.type_input.setText(material.type)
        
        self.balance_input = SpinBox()
        self.balance_input.setRange(0, 999999)
        self.balance_input.setValue(material.balance)
        
        self.price_input = SpinBox()
        self.price_input.setRange(1, 999999)
        self.price_input.setValue(material.price)
        self.price_input.setSuffix(" ₽")
        
        # Add fields to form
        fields_layout.addRow("Тип материала:", self.type_input)
        fields_layout.addRow("Остаток:", self.balance_input)
        fields_layout.addRow("Цена за единицу:", self.price_input)
        
        form_layout.addLayout(fields_layout)
        
        # Insert our form in place of content label
        self.textLayout.insertWidget(2, self.form_widget)
        
        # Update dialog size
        self.widget.setMinimumWidth(400)
        self.widget.setMinimumHeight(300)
        
        # Update button text
        self.yesButton.setText("Сохранить")
        self.cancelButton.setText("Отмена")
        
    def get_material_data(self):
        return {
            "type": self.type_input.text(),
            "balance": self.balance_input.value(),
            "price": self.price_input.value()
        }


class MaterialsInterface(ScrollArea):
    def __init__(self, user_data, parent=None):
        super().__init__(parent=parent)
        self.user_data = user_data
        self.material_controller = MaterialController()
        
        # Create widget and layout
        self.scroll_widget = QWidget()
        self.scroll_layout = QVBoxLayout(self.scroll_widget)
        self.scroll_layout.setSpacing(20)
        self.scroll_layout.setContentsMargins(30, 30, 30, 30)
        self.setWidget(self.scroll_widget)
        self.setWidgetResizable(True)
        
        # Set up UI
        self._setup_ui()
        
        # Load materials
        self.load_materials()
        
        # Connect signals
        self._connect_signals()
        
    def _setup_ui(self):
        # Header
        self.main_layout = QVBoxLayout(self)
        self.main_layout.setContentsMargins(30, 30, 30, 30)
        self.main_layout.setSpacing(20)
        
        # Create progress bar for loading
        self.progress_bar = IndeterminateProgressBar(self)
        self.progress_bar.setVisible(False)
        
        # Title and search
        header_layout = QHBoxLayout()
        
        # Title
        title_label = SubtitleLabel("Материалы")
        title_label.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        header_layout.addWidget(title_label)
        
        header_layout.addStretch()
        
        # Search
        self.search_box = SearchLineEdit(self)
        self.search_box.setPlaceholderText("Поиск по типу материала")
        self.search_box.setFixedWidth(300)
        self.search_box.textChanged.connect(self.filter_materials)
        header_layout.addWidget(self.search_box)
        
        # Add material button
        self.add_button = PrimaryPushButton("Добавить материал")
        self.add_button.setIcon(FluentIcon.ADD)
        self.add_button.clicked.connect(self.add_material)
        header_layout.addWidget(self.add_button)
        
        # Generate report button
        self.report_button = PushButton("Сформировать отчет")
        self.report_button.setIcon(FluentIcon.DOCUMENT)
        self.report_button.clicked.connect(self.generate_report)
        header_layout.addWidget(self.report_button)
        
        self.main_layout.addLayout(header_layout)
        
        # Create table
        self.materials_table = TableWidget(self)
        self.materials_table.setColumnCount(5)
        self.materials_table.setHorizontalHeaderLabels(["ID", "Тип материала", "Остаток", "Цена", "Действия"])
        self.materials_table.setColumnHidden(0, True)  # Hide ID column
        self.materials_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.materials_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self.materials_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.materials_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.materials_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        
        self.main_layout.addWidget(self.materials_table)
        
        # Add progress bar to layout
        self.main_layout.addWidget(self.progress_bar)
        
    def _connect_signals(self):
        # Connect to signal bus
        signalBus.database_error.connect(self.show_db_error)
        
    def load_materials(self):
        """Load materials from database"""
        # Show loading indicator
        self.progress_bar.setVisible(True)
        
        # Clear existing table
        self.materials_table.setRowCount(0)
        
        try:
            db = SessionLocal()
            
            # Get all materials
            materials = self.material_controller.get_all(db)
            
            # Set row count
            self.materials_table.setRowCount(len(materials))
            
            # Fill table with data
            for i, material in enumerate(materials):
                # ID (hidden)
                id_item = QTableWidgetItem(str(material.id))  # Ensure ID is converted to string
                self.materials_table.setItem(i, 0, id_item)
                
                # Type
                type_item = QTableWidgetItem(material.type)
                self.materials_table.setItem(i, 1, type_item)
                
                # Balance
                balance_item = QTableWidgetItem(str(material.balance))
                self.materials_table.setItem(i, 2, balance_item)
                
                # Price
                price_item = QTableWidgetItem(f"{material.price} ₽")
                self.materials_table.setItem(i, 3, price_item)
                
                # Actions
                actions_widget = QWidget()
                actions_layout = QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(0, 0, 0, 0)
                actions_layout.setSpacing(5)
                
                # Store material ID for use with button callbacks
                material_id = material.id
                material_obj = material
                
                # Edit button - fixed connection
                edit_button = PushButton("Редактировать")
                edit_button.setIcon(FluentIcon.EDIT)
                # Use a safer approach for button connections
                edit_button_callback = lambda checked=False, mat=material_obj: self.edit_material(mat)
                edit_button.clicked.connect(edit_button_callback)
                actions_layout.addWidget(edit_button)
                
                self.materials_table.setCellWidget(i, 4, actions_widget)
                
                # Make cells read-only
                for j in range(4):
                    item = self.materials_table.item(i, j)
                    if item:
                        item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                        
            # Show info if no materials
            if len(materials) == 0:
                InfoBar.info(
                    title="Нет материалов",
                    content="В базе данных нет материалов. Добавьте новый материал.",
                    parent=self,
                    duration=8000  # Increase duration to 8 seconds for better visibility
                )
            else:
                # If materials were loaded successfully, show a success message
                InfoBar.success(
                    title="Материалы загружены",
                    content=f"Загружено материалов: {len(materials)}",
                    parent=self,
                    duration=3000
                )
        except Exception as e:
            InfoBar.error(
                title="Ошибка загрузки материалов",
                content=str(e),
                parent=self,
                duration=8000  # Increase duration to 8 seconds for better error visibility
            )
        finally:
            db.close()  # Properly close the database session
            self.progress_bar.setVisible(False)  # Hide progress bar regardless of result
            
    def filter_materials(self, text):
        """Filter materials table based on search text"""
        text = text.lower()
        for i in range(self.materials_table.rowCount()):
            visible = False
            # Search in type column
            type_item = self.materials_table.item(i, 1)
            if type_item and text in type_item.text().lower():
                visible = True
                
            self.materials_table.setRowHidden(i, not visible)
            
    def add_material(self):
        dialog = AddMaterialDialog(self)
        if dialog.exec():
            material_data = dialog.get_material_data()
            
            try:
                material_create = MaterialCreate(**material_data)
            except Exception as e:
                InfoBar.error(
                    title="Ошибка валидации данных",
                    content=str(e),
                    parent=self
                )
                return
                
            try:
                db = SessionLocal()
                created_material = self.material_controller.create(db, material_create)
                
                if created_material:
                    InfoBar.success(
                        title="Материал добавлен",
                        content=f"Материал '{created_material.type}' успешно добавлен",
                        parent=self
                    )
                    # Reload materials
                    self.load_materials()
            except Exception as e:
                InfoBar.error(
                    title="Ошибка добавления материала",
                    content=str(e),
                    parent=self
                )
            finally:
                SessionLocal.remove()
                
    def edit_material(self, material):
        dialog = EditMaterialDialog(material, self)
        if dialog.exec():
            material_data = dialog.get_material_data()
            
            try:
                material_update = MaterialUpdate(**material_data)
            except Exception as e:
                InfoBar.error(
                    title="Ошибка валидации данных",
                    content=str(e),
                    parent=self
                )
                return
                
            try:
                db = SessionLocal()
                updated_material = self.material_controller.update(db, material.id, material_update)
                
                if updated_material:
                    InfoBar.success(
                        title="Материал обновлен",
                        content=f"Материал '{updated_material.type}' успешно обновлен",
                        parent=self
                    )
                    # Reload materials
                    self.load_materials()
            except Exception as e:
                InfoBar.error(
                    title="Ошибка обновления материала",
                    content=str(e),
                    parent=self
                )
            finally:
                SessionLocal.remove()
                
    def delete_material(self, material):
        # Confirm deletion
        confirmation = MessageBox(
            "Удаление материала",
            f"Вы уверены, что хотите удалить материал '{material.type}'?",
            self
        )
        
        if confirmation.exec():
            try:
                db = SessionLocal()
                deleted = self.material_controller.delete(db, material.id)
                
                if deleted:
                    InfoBar.success(
                        title="Материал удален",
                        content=f"Материал '{material.type}' успешно удален",
                        parent=self
                    )
                    # Reload materials
                    self.load_materials()
                else:
                    InfoBar.warning(
                        title="Ошибка удаления",
                        content=f"Не удалось удалить материал '{material.type}'",
                        parent=self
                    )
            except Exception as e:
                InfoBar.error(
                    title="Ошибка удаления материала",
                    content=str(e),
                    parent=self
                )
            finally:
                SessionLocal.remove()
                
    def generate_report(self):
        try:
            db = SessionLocal()
            materials = self.material_controller.get_all(db)
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Остатки материалов"
            
            # Add header
            ws.append(["Тип материала", "Остаток", "Цена за единицу", "Общая стоимость"])
            
            # Style header
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font
            
            # Add data
            total_value = 0
            if materials:
                for material in materials:
                    total_cost = material.balance * material.price
                    total_value += total_cost
                    ws.append([material.type, material.balance, f"{material.price} ₽", f"{total_cost} ₽"])
            else:
                # Add a row indicating no materials
                ws.append(["Нет материалов в базе данных", 0, "0 ₽", "0 ₽"])
                
            # Add total row
            ws.append(["", "", "Общая стоимость:", f"{total_value} ₽"])
            total_row = ws.max_row
            for cell in ws[total_row]:
                cell.font = Font(bold=True)
                
            # Auto-fit columns
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[col_letter].width = adjusted_width
                
            # Create directory if not exists
            reports_dir = os.path.join(os.path.expanduser("~"), "Terra Reports")
            os.makedirs(reports_dir, exist_ok=True)
            
            # Save file
            date_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            file_path = os.path.join(reports_dir, f"materials_report_{date_str}.xlsx")
            wb.save(file_path)
            
            # Show success message
            InfoBar.success(
                title="Отчет сформирован",
                content=f"Отчет сохранен в файл: {file_path}",
                parent=self
            )
        except Exception as e:
            InfoBar.error(
                title="Ошибка формирования отчета",
                content=str(e),
                parent=self
            )
        finally:
            SessionLocal.remove()
            
    @pyqtSlot(str)
    def show_db_error(self, message):
        # Show database error
        InfoBar.error(
            title="Ошибка базы данных",
            content=message,
            parent=self
        ) 